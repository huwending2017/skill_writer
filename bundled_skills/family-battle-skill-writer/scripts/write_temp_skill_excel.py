#!/usr/bin/env python
from __future__ import annotations

import argparse
import copy
import json
import re
import shutil
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Set, Tuple

from openpyxl import load_workbook  # type: ignore

try:
    import win32com.client  # type: ignore
except ImportError:  # pragma: no cover - optional legacy engine
    win32com = None  # type: ignore


SKILL_WORKBOOK_KEY = "skill_workbook"
WAR_WORKBOOK_KEY = "war_paper_workbook"
DEFAULT_MAX_LEVEL = 10
SHORT_MAX_LEVEL = 1
ACTION_LOG_RE = re.compile(r"^(?P<sheet>\w+): (?P<action>insert|update) row=(?P<row>\d+) key=(?P<key>.+)$")
DEDUPE_LOG_RE = re.compile(r"^(?P<sheet>\w+): dedupe keep_row=(?P<keep>\d+) remove_rows=(?P<rows>\[[^\]]*\]) key=(?P<key>.+)$")
REORDER_LOG_RE = re.compile(r"^(?P<sheet>\w+): reorder rows=(?P<rows>\d+-\d+) keys=(?P<keys>.+)$")


SHEET_SCHEMAS: Dict[str, Dict[str, Any]] = {
    "skill": {
        "workbook_key": SKILL_WORKBOOK_KEY,
        "sheet_name": "skill",
        "data_start_row": 4,
        "unique_column": 1,
        "field_to_col": {
            "key": 1,
            "id": 2,
            "name": 3,
            "desc": 4,
            "de_desc": 6,
            "quality": 7,
            "init_skill": 8,
            "type": 9,
            "skill_type": 10,
            "work_type": 11,
            "appoint_type": 12,
            "target": 13,
            "ready": 14,
            "cd": 15,
            "skill_lv": 16,
            "max_lv": 17,
            "coin_cost": 18,
            "chance": 19,
            "fit_arms": 20,
            "study_need": 21,
            "learn_times": 22,
            "special_script": 23,
            "special_param": 24,
            "person": 25,
            "inherit_hero": 26,
            "skill_up_num": 27,
            "star_up_num": 28,
            "icon": 30,
            "replay_skill_show": 31,
            "effect_clash": 32,
            "user_range": 33,
            "hide": 34,
            "owner_hero": 35,
            "season": 36,
            "hide_in_list": 37,
            "skill_analysis": 38,
        },
    },
    "skill_stage": {
        "workbook_key": SKILL_WORKBOOK_KEY,
        "sheet_name": "skill_stage",
        "data_start_row": 4,
        "unique_column": 1,
        "field_to_col": {
            "key": 1,
            "skill_id": 2,
            "desc": 3,
            "skill_level": 4,
            "stage": 5,
            "camp": 6,
            "target": 7,
            "sex": 8,
            "attr": 9,
            "script": 10,
            "param": 11,
            "chance": 12,
            "attack_per": 13,
            "effect_id": 14,
            "hit_effect_id": 15,
            "action_id": 16,
            "show_range": 17,
        },
    },
    "buff": {
        "workbook_key": SKILL_WORKBOOK_KEY,
        "sheet_name": "buff",
        "data_start_row": 4,
        "unique_column": 1,
        "field_to_col": {
            "key": 1,
            "id": 2,
            "name": 3,
            "desc": 4,
            "time_add": 5,
            "level": 6,
            "type": 7,
            "is_dispel": 8,
            "type_id": 9,
            "is_dead_clear": 10,
            "ready": 11,
            "time": 12,
            "add_type": 13,
            "add_max": 14,
            "update_life": 15,
            "chance": 16,
            "script": 17,
            "param": 18,
            "effect_pos": 19,
            "effect_id": 20,
            "action_id": 21,
            "is_loop": 22,
            "icon_id": 23,
            "buff_desc": 24,
            "buff_character": 25,
            "attack_per": 26,
            "effect_id_2": 27,
            "replay_skill_show": 28,
            "effect_type": 29,
        },
    },
    "war_paper": {
        "workbook_key": WAR_WORKBOOK_KEY,
        "sheet_name": "war_paper",
        "data_start_row": 4,
        "unique_column": 2,
        "field_to_col": {
            "ID": 1,
            "name": 2,
            "is_show": 3,
            "type": 4,
            "desc1": 5,
            "data_order1": 6,
            "cost_time": 7,
            "action_type": 8,
            "key_param": 9,
            "effect_desc": 10,
            "beizhu2": 11,
            "param1": 12,
            "param2": 13,
            "param3": 14,
            "param4": 15,
            "param5": 16,
            "param6": 17,
            "param7": 18,
            "param8": 19,
        },
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Write temporary battle skill payload rows back into Excel workbooks."
    )
    parser.add_argument("--payload", required=True, help="Path to temp_excel_payload.json")
    parser.add_argument("--skill-xlsx", help="Override path to J_技能表_skill.xlsx")
    parser.add_argument("--war-xlsx", help="Override path to Z_战报表.xlsx")
    parser.add_argument(
        "--copy-to",
        help="Copy target workbooks into this folder first, then write into the copies.",
    )
    parser.add_argument(
        "--backup-dir",
        help="Create timestamped backups here before writing to the real workbook paths.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview row insert/update actions without saving the workbook.",
    )
    parser.add_argument(
        "--dedupe-existing",
        action="store_true",
        help="Delete duplicate existing rows that match the same payload row after upsert.",
    )
    parser.add_argument(
        "--engine",
        choices=("openpyxl", "excel"),
        default="openpyxl",
        help="Excel write engine. openpyxl is much faster; excel keeps the old COM path.",
    )
    return parser.parse_args()


def load_payload(path: Path) -> Dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if "rows" not in data or not isinstance(data["rows"], dict):
        raise ValueError("payload must contain a top-level 'rows' object")
    suspicious_fields = find_suspicious_question_mark_fields(data)
    if suspicious_fields:
        details = "\n".join(f"  - {item}" for item in suspicious_fields[:20])
        raise ValueError(
            "payload contains suspicious question-mark mojibake in user-visible fields. "
            "Fix temp_excel_payload.json before writing Excel.\n" + details
        )
    return data


USER_VISIBLE_PAYLOAD_FIELDS = {
    "name",
    "desc",
    "de_desc",
    "buff_desc",
    "desc1",
    "effect_desc",
    "beizhu2",
    "param1",
    "param2",
    "param3",
    "param4",
    "param5",
    "param6",
    "param7",
    "param8",
}


def find_suspicious_question_mark_fields(payload: Dict[str, Any], limit: int = 20) -> List[str]:
    rows = payload.get("rows")
    if not isinstance(rows, dict):
        return []

    findings: List[str] = []
    for sheet_name, sheet_rows in rows.items():
        if not isinstance(sheet_rows, list):
            continue
        for index, row in enumerate(sheet_rows, start=1):
            if not isinstance(row, dict):
                continue
            row_key = row.get("key") or row.get("name") or row.get("id") or index
            for field in USER_VISIBLE_PAYLOAD_FIELDS:
                value = row.get(field)
                if not isinstance(value, str):
                    continue
                if "??" in value or value.count("?") >= 3:
                    findings.append(f"{sheet_name}[{index}] key={row_key} field={field} value={value[:80]}")
                    if len(findings) >= limit:
                        return findings
    return findings


def now_tag() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def ensure_path(value: str | None, label: str) -> Path:
    if not value:
        raise ValueError(f"missing workbook path for {label}")
    return Path(value).expanduser().resolve()


def required_workbook_keys(payload: Dict[str, Any]) -> Set[str]:
    required: Set[str] = set()
    for sheet_name, rows in payload.get("rows", {}).items():
        if not rows:
            continue
        if sheet_name not in SHEET_SCHEMAS:
            raise ValueError(f"unsupported sheet name: {sheet_name}")
        required.add(SHEET_SCHEMAS[sheet_name]["workbook_key"])
    return required


def resolve_workbook_targets(
    args: argparse.Namespace,
    payload: Dict[str, Any],
    required_keys: Set[str] | None = None,
) -> Dict[str, Path]:
    targets = payload.get("targets", {})
    path_values = {
        SKILL_WORKBOOK_KEY: args.skill_xlsx or targets.get(SKILL_WORKBOOK_KEY),
        WAR_WORKBOOK_KEY: args.war_xlsx or targets.get(WAR_WORKBOOK_KEY),
    }
    if required_keys is None:
        required_keys = set(path_values)

    resolved: Dict[str, Path] = {}
    for key in required_keys:
        resolved[key] = ensure_path(path_values.get(key), key)
    for path in resolved.values():
        if not path.exists():
            raise FileNotFoundError(path)
    return resolved


def maybe_copy_targets(targets: Dict[str, Path], copy_to: str | None) -> Dict[str, Path]:
    if not copy_to:
        return targets
    out_dir = Path(copy_to).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    copied: Dict[str, Path] = {}
    for key, src in targets.items():
        dst = out_dir / src.name
        if dst.exists():
            suffix = time.strftime("%Y%m%d_%H%M%S")
            dst = out_dir / f"{src.stem}_{suffix}{src.suffix}"
        shutil.copy2(src, dst)
        copied[key] = dst
    return copied


def maybe_backup_targets(targets: Dict[str, Path], backup_dir: str | None, dry_run: bool) -> None:
    if not backup_dir or dry_run:
        return
    out_dir = Path(backup_dir).expanduser().resolve() / now_tag()
    out_dir.mkdir(parents=True, exist_ok=True)
    for src in targets.values():
        shutil.copy2(src, out_dir / src.name)


def derive_skill_key(row: Dict[str, Any]) -> str:
    return f"{row['id']}_{row['skill_lv']}"


def derive_stage_key(row: Dict[str, Any]) -> str:
    stage = row.get("stage", row.get("id"))
    return f"{row['skill_id']}_{stage}_{row['skill_level']}"


def derive_buff_key(row: Dict[str, Any]) -> str:
    return f"{row['id']}_{row['level']}"


def normalize_row(sheet_name: str, raw_row: Dict[str, Any]) -> Dict[str, Any]:
    row = dict(raw_row)
    if sheet_name == "skill":
        row["key"] = derive_skill_key(row)
    elif sheet_name == "skill_stage":
        if "stage" not in row and "id" in row:
            row["stage"] = row["id"]
        row["key"] = derive_stage_key(row)
    elif sheet_name == "buff":
        if "update_life" not in row and "opportunity" in row:
            row["update_life"] = row["opportunity"]
        row["key"] = derive_buff_key(row)
    elif sheet_name == "war_paper":
        if "record_id" in row and "ID" not in row:
            row["ID"] = row["record_id"]
        if "record_name" in row and "name" not in row:
            row["name"] = row["record_name"]
        if "id" in row and "ID" not in row:
            row["ID"] = row["id"]
    return row


def _to_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


MAX_LEVEL_FIELDS = ("max_lv", "max_level", "skill_max_lv", "skill_max_level", "level_max")
CATEGORY_FIELDS = ("skill_category", "category", "source_type", "skill_source", "skill_kind", "kind")


def _explicit_max_levels(rows: Iterable[Dict[str, Any]]) -> List[int]:
    levels: List[int] = []
    for row in rows:
        for field in MAX_LEVEL_FIELDS:
            if row.get(field) is None:
                continue
            level = _to_int(row.get(field), -1)
            if level >= 0:
                levels.append(level)
    return levels


def _category_default_max_level(rows: Iterable[Dict[str, Any]], default: int = DEFAULT_MAX_LEVEL) -> int:
    for row in rows:
        for field in CATEGORY_FIELDS:
            value = row.get(field)
            if not isinstance(value, str):
                continue
            if "兵书" in value or "装备" in value:
                return SHORT_MAX_LEVEL
            if "自带" in value:
                return DEFAULT_MAX_LEVEL
    return default


def _expand_skill_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[int, List[Dict[str, Any]]] = {}
    for row in rows:
        skill_id = _to_int(row.get("id"), -1)
        grouped.setdefault(skill_id, []).append(dict(row))

    expanded: List[Dict[str, Any]] = []
    for skill_id in sorted(grouped):
        group = grouped[skill_id]
        template = max(group, key=lambda item: _to_int(item.get("skill_lv"), 0))
        explicit_max_levels = _explicit_max_levels(group)
        if explicit_max_levels:
            max_lv = max(explicit_max_levels)
        else:
            max_lv = max(
                _category_default_max_level(group),
                max(_to_int(item.get("skill_lv"), 0) for item in group),
            )
        by_level = {_to_int(item.get("skill_lv"), 0): dict(item) for item in group}
        for level in range(0, max_lv + 1):
            row = dict(by_level.get(level, template))
            row["id"] = skill_id
            row["skill_lv"] = level
            row["max_lv"] = max_lv
            expanded.append(normalize_row("skill", row))
    return expanded


def _expand_stage_rows(rows: List[Dict[str, Any]], skill_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    skill_max_levels = {
        _to_int(row.get("id"), -1): _to_int(row.get("max_lv"), DEFAULT_MAX_LEVEL)
        for row in skill_rows
    }
    grouped: Dict[Tuple[int, int], List[Dict[str, Any]]] = {}
    for row in rows:
        skill_id = _to_int(row.get("skill_id"), -1)
        stage = _to_int(row.get("stage", row.get("id")), 0)
        grouped.setdefault((skill_id, stage), []).append(dict(row))

    expanded: List[Dict[str, Any]] = []
    for (skill_id, stage) in sorted(grouped):
        group = grouped[(skill_id, stage)]
        template = max(group, key=lambda item: _to_int(item.get("skill_level"), 0))
        max_lv = skill_max_levels.get(
            skill_id,
            max(_to_int(item.get("skill_level"), DEFAULT_MAX_LEVEL) for item in group),
        )
        by_level = {_to_int(item.get("skill_level"), 0): dict(item) for item in group}
        for level in range(0, max_lv + 1):
            row = dict(by_level.get(level, template))
            row["skill_id"] = skill_id
            row["stage"] = stage
            row["skill_level"] = level
            expanded.append(normalize_row("skill_stage", row))
    return expanded


def _expand_buff_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[int, List[Dict[str, Any]]] = {}
    for row in rows:
        buff_id = _to_int(row.get("id"), -1)
        grouped.setdefault(buff_id, []).append(dict(row))

    expanded: List[Dict[str, Any]] = []
    for buff_id in sorted(grouped):
        group = grouped[buff_id]
        template = max(group, key=lambda item: _to_int(item.get("level"), 0))
        explicit_max_levels = _explicit_max_levels(group)
        if explicit_max_levels:
            max_lv = max(explicit_max_levels)
        else:
            max_lv = max(_to_int(item.get("level"), DEFAULT_MAX_LEVEL) for item in group)
            max_lv = max(max_lv, _category_default_max_level(group))
        by_level = {_to_int(item.get("level"), 0): dict(item) for item in group}
        for level in range(0, max_lv + 1):
            row = dict(by_level.get(level, template))
            row["id"] = buff_id
            row["level"] = level
            row["max_lv"] = max_lv
            expanded.append(normalize_row("buff", row))
    return expanded


def expand_payload_rows(payload: Dict[str, Any]) -> Dict[str, Any]:
    rows = payload.get("rows", {})
    skill_rows = [normalize_row("skill", row) for row in rows.get("skill", [])]
    skill_rows = _expand_skill_rows(skill_rows)
    stage_rows = [normalize_row("skill_stage", row) for row in rows.get("skill_stage", [])]
    stage_rows = _expand_stage_rows(stage_rows, skill_rows)
    buff_rows = [normalize_row("buff", row) for row in rows.get("buff", [])]
    buff_rows = _expand_buff_rows(buff_rows)
    war_rows = [normalize_row("war_paper", row) for row in rows.get("war_paper", [])]

    expanded_payload = dict(payload)
    expanded_payload["rows"] = {
        "skill": skill_rows,
        "skill_stage": stage_rows,
        "buff": buff_rows,
        "war_paper": war_rows,
    }
    return expanded_payload


def stringify_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, (int, str)):
        return str(value)
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def normalize_match_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if not text:
        return ""
    try:
        number = float(text)
        return str(int(number)) if number.is_integer() else text
    except ValueError:
        return text


def stringify_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        if not value:
            return ""
        if any(isinstance(item, (list, tuple, dict)) for item in value):
            parts = []
            for item in value:
                if isinstance(item, (list, tuple)):
                    parts.append(",".join(stringify_scalar(v) for v in item))
                else:
                    parts.append(stringify_scalar(item))
            return "|".join(parts)
        return ",".join(stringify_scalar(item) for item in value)
    return stringify_scalar(value)


def get_unique_value(sheet_name: str, row: Dict[str, Any]) -> str:
    if sheet_name == "war_paper":
        if "name" in row:
            return str(row["name"])
        if "ID" in row:
            return str(row["ID"])
        raise ValueError("war_paper row must contain name or ID")
    return str(row["key"])


def get_match_signatures(sheet_name: str, row: Dict[str, Any]) -> List[Tuple[str, str]]:
    signatures: List[Tuple[str, str]] = []

    def add(label: str, value: Any) -> None:
        text = normalize_match_value(value)
        if text:
            signatures.append((label, text))

    def add_joined(label: str, *values: Any) -> None:
        parts = [normalize_match_value(value) for value in values]
        if not parts or any(not part for part in parts):
            return
        signatures.append((label, "|".join(parts)))

    if sheet_name == "skill":
        add("key", row.get("key"))
        add_joined("id_lv", row.get("id"), row.get("skill_lv"))
    elif sheet_name == "skill_stage":
        add("key", row.get("key"))
        add_joined("id_stage_lv", row.get("skill_id"), row.get("stage"), row.get("skill_level"))
    elif sheet_name == "buff":
        add("key", row.get("key"))
        add_joined("id_lv", row.get("id"), row.get("level"))
    elif sheet_name == "war_paper":
        add("name", row.get("name"))
        add("id", row.get("ID"))
    else:
        add("key", row.get("key"))
    return signatures


def signature_read_max_col(sheet_name: str) -> int:
    if sheet_name == "skill":
        return 16
    if sheet_name == "skill_stage":
        return 5
    if sheet_name == "buff":
        return 6
    if sheet_name == "war_paper":
        return 2
    return SHEET_SCHEMAS[sheet_name]["unique_column"]


def get_last_row(ws: Any, unique_column: int, data_start_row: int) -> int:
    xl_up = -4162
    last_row = ws.Cells(ws.Rows.Count, unique_column).End(xl_up).Row
    return max(last_row, data_start_row - 1)


def get_existing_row_signatures(ws: Any, sheet_name: str, row_idx: int) -> List[Tuple[str, str]]:
    if sheet_name == "skill":
        return get_match_signatures(
            sheet_name,
            {
                "key": ws.Cells(row_idx, 1).Value,
                "id": ws.Cells(row_idx, 2).Value,
                "skill_lv": ws.Cells(row_idx, 16).Value,
            },
        )
    if sheet_name == "skill_stage":
        return get_match_signatures(
            sheet_name,
            {
                "key": ws.Cells(row_idx, 1).Value,
                "skill_id": ws.Cells(row_idx, 2).Value,
                "skill_level": ws.Cells(row_idx, 4).Value,
                "stage": ws.Cells(row_idx, 5).Value,
            },
        )
    if sheet_name == "buff":
        return get_match_signatures(
            sheet_name,
            {
                "key": ws.Cells(row_idx, 1).Value,
                "id": ws.Cells(row_idx, 2).Value,
                "level": ws.Cells(row_idx, 6).Value,
            },
        )
    if sheet_name == "war_paper":
        return get_match_signatures(
            sheet_name,
            {
                "ID": ws.Cells(row_idx, 1).Value,
                "name": ws.Cells(row_idx, 2).Value,
            },
        )
    return []


def normalize_range_values(values: Any, expected_cols: int) -> List[Tuple[Any, ...]]:
    if values is None:
        return []
    if not isinstance(values, tuple):
        return [(values,)]
    if not values:
        return []
    first = values[0]
    if isinstance(first, tuple):
        return [tuple(row) for row in values]
    if expected_cols == 1:
        return [(value,) for value in values]
    return [tuple(values)]


def read_sheet_values(ws: Any, start_row: int, last_row: int, max_col: int) -> List[Tuple[Any, ...]]:
    if last_row < start_row:
        return []
    values = ws.Range(ws.Cells(start_row, 1), ws.Cells(last_row, max_col)).Value
    return normalize_range_values(values, max_col)


def value_at(row_values: Sequence[Any], col: int) -> Any:
    index = col - 1
    if index < 0 or index >= len(row_values):
        return None
    return row_values[index]


def get_existing_row_signatures_from_values(
    sheet_name: str,
    row_values: Sequence[Any],
) -> List[Tuple[str, str]]:
    if sheet_name == "skill":
        return get_match_signatures(
            sheet_name,
            {
                "key": value_at(row_values, 1),
                "id": value_at(row_values, 2),
                "skill_lv": value_at(row_values, 16),
            },
        )
    if sheet_name == "skill_stage":
        return get_match_signatures(
            sheet_name,
            {
                "key": value_at(row_values, 1),
                "skill_id": value_at(row_values, 2),
                "skill_level": value_at(row_values, 4),
                "stage": value_at(row_values, 5),
            },
        )
    if sheet_name == "buff":
        return get_match_signatures(
            sheet_name,
            {
                "key": value_at(row_values, 1),
                "id": value_at(row_values, 2),
                "level": value_at(row_values, 6),
            },
        )
    if sheet_name == "war_paper":
        return get_match_signatures(
            sheet_name,
            {
                "ID": value_at(row_values, 1),
                "name": value_at(row_values, 2),
            },
        )
    return []


def build_existing_signature_map(ws: Any, sheet_name: str, data_start_row: int) -> Dict[Tuple[str, str], List[int]]:
    unique_column = SHEET_SCHEMAS[sheet_name]["unique_column"]
    last_row = get_last_row(ws, unique_column, data_start_row)
    max_col = signature_read_max_col(sheet_name)
    sheet_values = read_sheet_values(ws, data_start_row, last_row, max_col)
    signature_map: Dict[Tuple[str, str], List[int]] = {}
    for offset, row_values in enumerate(sheet_values):
        row_idx = data_start_row + offset
        for signature in get_existing_row_signatures_from_values(sheet_name, row_values):
            rows = signature_map.setdefault(signature, [])
            if row_idx not in rows:
                rows.append(row_idx)
    return signature_map


def build_existing_sheet_state(
    ws: Any,
    sheet_name: str,
    data_start_row: int,
) -> Tuple[Dict[Tuple[str, str], List[int]], Dict[int, List[Any]], int]:
    schema = SHEET_SCHEMAS[sheet_name]
    last_row = get_last_row(ws, schema["unique_column"], data_start_row)
    max_col = signature_read_max_col(sheet_name)
    sheet_values = read_sheet_values(ws, data_start_row, last_row, max_col)
    signature_map: Dict[Tuple[str, str], List[int]] = {}
    row_values_by_index: Dict[int, List[Any]] = {}

    for offset, row_values in enumerate(sheet_values):
        row_idx = data_start_row + offset
        normalized_values = list(row_values)
        if len(normalized_values) < max_col:
            normalized_values.extend([""] * (max_col - len(normalized_values)))
        row_values_by_index[row_idx] = normalized_values
        for signature in get_existing_row_signatures_from_values(sheet_name, normalized_values):
            rows = signature_map.setdefault(signature, [])
            if row_idx not in rows:
                rows.append(row_idx)

    return signature_map, row_values_by_index, last_row


def register_row_signatures(
    signature_map: Dict[Tuple[str, str], List[int]],
    signatures: Iterable[Tuple[str, str]],
    row_idx: int,
) -> None:
    for signature in signatures:
        rows = signature_map.setdefault(signature, [])
        if row_idx not in rows:
            rows.append(row_idx)


def delete_rows_desc(ws: Any, row_indexes: Iterable[int], dry_run: bool) -> List[str]:
    deleted_logs: List[str] = []
    for row_idx in sorted(set(row_indexes), reverse=True):
        deleted_logs.append(f"delete duplicate row={row_idx}")
        if not dry_run:
            ws.Rows(row_idx).Delete()
    return deleted_logs


def write_row_values(ws: Any, row_idx: int, values: Sequence[Any]) -> None:
    if not values:
        return
    ws.Range(ws.Cells(row_idx, 1), ws.Cells(row_idx, len(values))).Value = (tuple(values),)


def build_openpyxl_sheet_state(
    ws: Any,
    sheet_name: str,
    data_start_row: int,
) -> Tuple[Dict[Tuple[str, str], List[int]], int]:
    schema = SHEET_SCHEMAS[sheet_name]
    unique_column = schema["unique_column"]
    max_col = signature_read_max_col(sheet_name)
    signature_map: Dict[Tuple[str, str], List[int]] = {}
    last_row = data_start_row - 1

    for row_idx, row_values in enumerate(
        ws.iter_rows(min_row=data_start_row, max_col=max_col, values_only=True),
        start=data_start_row,
    ):
        normalized_values = list(row_values)
        if len(normalized_values) < max_col:
            normalized_values.extend([""] * (max_col - len(normalized_values)))
        if normalize_match_value(value_at(normalized_values, unique_column)):
            last_row = row_idx
        for signature in get_existing_row_signatures_from_values(sheet_name, normalized_values):
            rows = signature_map.setdefault(signature, [])
            if row_idx not in rows:
                rows.append(row_idx)

    return signature_map, last_row


def delete_rows_desc_openpyxl(ws: Any, row_indexes: Iterable[int], dry_run: bool) -> List[str]:
    deleted_logs: List[str] = []
    for row_idx in sorted(set(row_indexes), reverse=True):
        deleted_logs.append(f"delete duplicate row={row_idx}")
        if not dry_run:
            ws.delete_rows(row_idx, 1)
    return deleted_logs


def write_openpyxl_row_values(ws: Any, row_idx: int, values: Sequence[Any]) -> None:
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx).value = value


def write_openpyxl_row_fields(
    ws: Any,
    row_idx: int,
    field_to_col: Dict[str, int],
    row: Dict[str, Any],
) -> None:
    for field, col in field_to_col.items():
        if field not in row:
            continue
        ws.cell(row=row_idx, column=col).value = stringify_value(row[field])


def xlsx_ns(tag: str) -> str:
    return f"{{http://schemas.openxmlformats.org/spreadsheetml/2006/main}}{tag}"


def rel_ns(tag: str) -> str:
    return f"{{http://schemas.openxmlformats.org/package/2006/relationships}}{tag}"


def xml_fromstring(raw: bytes) -> Any:
    return ET.fromstring(raw)


def xml_iterparse(source: Any) -> Any:
    return ET.iterparse(source, events=("end",))


def col_index_from_ref(cell_ref: str) -> int:
    value = 0
    for char in cell_ref:
        if not char.isalpha():
            break
        value = value * 26 + (ord(char.upper()) - ord("A") + 1)
    return value


def parse_shared_strings(archive: zipfile.ZipFile) -> List[str]:
    try:
        raw = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    shared: List[str] = []
    root = xml_fromstring(raw)
    for si in root.findall(xlsx_ns("si")):
        parts: List[str] = []
        for text_node in si.iter(xlsx_ns("t")):
            parts.append(text_node.text or "")
        shared.append("".join(parts))
    return shared


def read_cell_text(cell: ET.Element, shared_strings: Sequence[str]) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "s":
        value_node = cell.find(xlsx_ns("v"))
        if value_node is None or value_node.text is None:
            return ""
        try:
            return shared_strings[int(value_node.text)]
        except (ValueError, IndexError):
            return ""
    if cell_type == "inlineStr":
        parts = [text_node.text or "" for text_node in cell.iter(xlsx_ns("t"))]
        return "".join(parts)
    value_node = cell.find(xlsx_ns("v"))
    return "" if value_node is None or value_node.text is None else value_node.text


def workbook_sheet_paths(archive: zipfile.ZipFile) -> Dict[str, str]:
    workbook_root = xml_fromstring(archive.read("xl/workbook.xml"))
    rels_root = xml_fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets: Dict[str, str] = {}
    for rel in rels_root.findall(rel_ns("Relationship")):
        rel_id = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if not rel_id or not target:
            continue
        target = target.lstrip("/")
        rel_targets[rel_id] = target if target.startswith("xl/") else f"xl/{target}"

    sheet_paths: Dict[str, str] = {}
    for sheet in workbook_root.findall(f"{xlsx_ns('sheets')}/{xlsx_ns('sheet')}"):
        sheet_name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
        if sheet_name and rel_id in rel_targets:
            sheet_paths[sheet_name] = rel_targets[rel_id]
    return sheet_paths


def build_xlsx_xml_sheet_state(
    workbook_path: Path,
    sheet_name: str,
    schema_sheet_name: str,
) -> Tuple[Dict[Tuple[str, str], List[int]], int]:
    with zipfile.ZipFile(workbook_path, "r") as archive:
        shared_strings = parse_shared_strings(archive)
        sheet_paths = workbook_sheet_paths(archive)
        return build_xlsx_xml_sheet_state_from_archive(
            archive,
            shared_strings,
            sheet_paths,
            sheet_name,
            schema_sheet_name,
        )


def build_xlsx_xml_sheet_state_from_archive(
    archive: zipfile.ZipFile,
    shared_strings: Sequence[str],
    sheet_paths: Dict[str, str],
    sheet_name: str,
    schema_sheet_name: str,
) -> Tuple[Dict[Tuple[str, str], List[int]], int]:
    data_start_row = SHEET_SCHEMAS[schema_sheet_name]["data_start_row"]
    unique_column = SHEET_SCHEMAS[schema_sheet_name]["unique_column"]
    max_col = signature_read_max_col(schema_sheet_name)
    signature_map: Dict[Tuple[str, str], List[int]] = {}
    last_row = data_start_row - 1

    sheet_path = sheet_paths.get(sheet_name)
    if not sheet_path:
        raise ValueError(f"sheet not found: {sheet_name}")

    with archive.open(sheet_path) as sheet_file:
        for event, elem in xml_iterparse(sheet_file):
            if elem.tag != xlsx_ns("row"):
                continue
            row_idx = _to_int(elem.attrib.get("r"), 0)
            if row_idx < data_start_row:
                elem.clear()
                continue

            values: List[Any] = [""] * max_col
            has_unique = False
            for cell in elem.findall(xlsx_ns("c")):
                cell_ref = cell.attrib.get("r", "")
                col_idx = col_index_from_ref(cell_ref)
                if col_idx <= 0 or col_idx > max_col:
                    continue
                text = read_cell_text(cell, shared_strings)
                values[col_idx - 1] = text
                if col_idx == unique_column and normalize_match_value(text):
                    has_unique = True
            if has_unique:
                last_row = row_idx
            for signature in get_existing_row_signatures_from_values(schema_sheet_name, values):
                rows = signature_map.setdefault(signature, [])
                if row_idx not in rows:
                    rows.append(row_idx)
            elem.clear()

    return signature_map, last_row


def plan_sheet_rows_from_state(
    sheet_name: str,
    rows: Iterable[Dict[str, Any]],
    signature_map: Dict[Tuple[str, str], List[int]],
    last_row: int,
    dedupe_existing: bool,
) -> List[str]:
    logs: List[str] = []
    rows_to_delete: Set[int] = set()

    for raw_row in rows:
        row = normalize_row(sheet_name, raw_row)
        unique_value = get_unique_value(sheet_name, row)
        row_signatures = get_match_signatures(sheet_name, row)
        matched_rows: Set[int] = set()
        for signature in row_signatures:
            matched_rows.update(signature_map.get(signature, []))

        row_idx = min(matched_rows) if matched_rows else None
        action = "update"
        if row_idx is None:
            row_idx = last_row + 1
            last_row = row_idx
            action = "insert"
        elif len(matched_rows) > 1:
            duplicate_rows = sorted(matched_rows - {row_idx})
            if dedupe_existing:
                rows_to_delete.update(duplicate_rows)
                logs.append(
                    f"{sheet_name}: dedupe keep_row={row_idx} remove_rows={duplicate_rows} key={unique_value}"
                )
            else:
                logs.append(
                    f"{sheet_name}: duplicate_detected keep_row={row_idx} duplicate_rows={duplicate_rows} key={unique_value}"
                )

        logs.append(f"{sheet_name}: {action} row={row_idx} key={unique_value}")
        register_row_signatures(signature_map, row_signatures, row_idx)

    for row_idx in sorted(rows_to_delete, reverse=True):
        logs.append(f"{sheet_name}: delete duplicate row={row_idx}")
    return logs


def dry_run_with_xlsx_xml(
    payload: Dict[str, Any],
    write_targets: Dict[str, Path],
    dedupe_existing: bool,
    timings: List[Tuple[str, float]],
) -> List[str]:
    all_logs: List[str] = []
    sheets_by_workbook: Dict[str, List[Tuple[str, List[Dict[str, Any]]]]] = {}
    for sheet_name, rows in payload["rows"].items():
        if sheet_name not in SHEET_SCHEMAS:
            raise ValueError(f"unsupported sheet name: {sheet_name}")
        if rows:
            workbook_key = SHEET_SCHEMAS[sheet_name]["workbook_key"]
            sheets_by_workbook.setdefault(workbook_key, []).append((sheet_name, rows))

    for workbook_key, sheet_items in sheets_by_workbook.items():
        workbook_path = write_targets[workbook_key]
        phase_started_at = time.perf_counter()
        with zipfile.ZipFile(workbook_path, "r") as archive:
            shared_strings = parse_shared_strings(archive)
            sheet_paths = workbook_sheet_paths(archive)
            timings.append((f"xml_open_{workbook_key}", time.perf_counter() - phase_started_at))
            for sheet_name, rows in sheet_items:
                schema = SHEET_SCHEMAS[sheet_name]
                phase_started_at = time.perf_counter()
                signature_map, last_row = build_xlsx_xml_sheet_state_from_archive(
                    archive,
                    shared_strings,
                    sheet_paths,
                    schema["sheet_name"],
                    sheet_name,
                )
                logs = plan_sheet_rows_from_state(sheet_name, rows, signature_map, last_row, dedupe_existing)
                timings.append((f"sheet_{sheet_name}", time.perf_counter() - phase_started_at))
                all_logs.extend(logs)
    return all_logs


def write_sheet_rows_openpyxl(
    ws: Any,
    sheet_name: str,
    rows: Iterable[Dict[str, Any]],
    dry_run: bool,
    dedupe_existing: bool,
) -> List[str]:
    schema = SHEET_SCHEMAS[sheet_name]
    data_start_row = schema["data_start_row"]
    field_to_col = schema["field_to_col"]
    signature_map, last_row = build_openpyxl_sheet_state(ws, sheet_name, data_start_row)
    logs: List[str] = []
    rows_to_delete: Set[int] = set()

    for raw_row in rows:
        row = normalize_row(sheet_name, raw_row)
        unique_value = get_unique_value(sheet_name, row)
        row_signatures = get_match_signatures(sheet_name, row)
        matched_rows: Set[int] = set()
        for signature in row_signatures:
            matched_rows.update(signature_map.get(signature, []))

        row_idx = min(matched_rows) if matched_rows else None
        action = "update"
        if row_idx is None:
            row_idx = last_row + 1
            last_row = row_idx
            action = "insert"
        elif len(matched_rows) > 1:
            duplicate_rows = sorted(matched_rows - {row_idx})
            if dedupe_existing:
                rows_to_delete.update(duplicate_rows)
                logs.append(
                    f"{sheet_name}: dedupe keep_row={row_idx} remove_rows={duplicate_rows} key={unique_value}"
                )
            else:
                logs.append(
                    f"{sheet_name}: duplicate_detected keep_row={row_idx} duplicate_rows={duplicate_rows} key={unique_value}"
                )

        logs.append(f"{sheet_name}: {action} row={row_idx} key={unique_value}")
        register_row_signatures(signature_map, row_signatures, row_idx)
        if dry_run:
            continue

        write_openpyxl_row_fields(ws, row_idx, field_to_col, row)

    if rows_to_delete:
        for log in delete_rows_desc_openpyxl(ws, rows_to_delete, dry_run):
            logs.append(f"{sheet_name}: {log}")
    return logs


def _openpyxl_capture_row(ws: Any, row_idx: int, max_col: int) -> Dict[str, Any]:
    return {
        "values": [ws.cell(row_idx, col).value for col in range(1, max_col + 1)],
        "styles": [copy.copy(ws.cell(row_idx, col)._style) for col in range(1, max_col + 1)],
        "number_formats": [ws.cell(row_idx, col).number_format for col in range(1, max_col + 1)],
        "fills": [copy.copy(ws.cell(row_idx, col).fill) for col in range(1, max_col + 1)],
        "fonts": [copy.copy(ws.cell(row_idx, col).font) for col in range(1, max_col + 1)],
        "borders": [copy.copy(ws.cell(row_idx, col).border) for col in range(1, max_col + 1)],
        "alignments": [copy.copy(ws.cell(row_idx, col).alignment) for col in range(1, max_col + 1)],
        "protections": [copy.copy(ws.cell(row_idx, col).protection) for col in range(1, max_col + 1)],
        "height": ws.row_dimensions[row_idx].height,
    }


def _openpyxl_restore_row(ws: Any, row_idx: int, saved: Dict[str, Any]) -> None:
    ws.row_dimensions[row_idx].height = saved["height"]
    for col, value in enumerate(saved["values"], start=1):
        cell = ws.cell(row_idx, col)
        cell.value = value
        cell._style = copy.copy(saved["styles"][col - 1])
        cell.number_format = saved["number_formats"][col - 1]
        cell.fill = copy.copy(saved["fills"][col - 1])
        cell.font = copy.copy(saved["fonts"][col - 1])
        cell.border = copy.copy(saved["borders"][col - 1])
        cell.alignment = copy.copy(saved["alignments"][col - 1])
        cell.protection = copy.copy(saved["protections"][col - 1])


def _payload_order_groups(sheet_name: str, rows: Iterable[Dict[str, Any]]) -> List[Tuple[Tuple[Any, ...], List[str]]]:
    if sheet_name not in {"skill", "skill_stage", "buff"}:
        return []

    grouped: Dict[Tuple[Any, ...], List[str]] = {}
    group_order: List[Tuple[Any, ...]] = []
    for raw_row in rows:
        row = normalize_row(sheet_name, raw_row)
        if sheet_name == "skill":
            group_key = (row.get("id"),)
        elif sheet_name == "skill_stage":
            group_key = (row.get("skill_id"),)
        else:
            group_key = (row.get("id"),)
        unique_value = str(get_unique_value(sheet_name, row))
        if group_key not in grouped:
            grouped[group_key] = []
            group_order.append(group_key)
        if unique_value not in grouped[group_key]:
            grouped[group_key].append(unique_value)

    return [(key, grouped[key]) for key in group_order if len(grouped[key]) > 1]


def _payload_group_anchor(
    ws: Any,
    sheet_name: str,
    group_key: Tuple[Any, ...],
    moving_rows: Set[int],
) -> int | None:
    if sheet_name not in {"skill", "skill_stage", "buff"}:
        return None

    schema = SHEET_SCHEMAS[sheet_name]
    field_to_col = schema["field_to_col"]
    id_col = field_to_col["skill_id"] if sheet_name == "skill_stage" else field_to_col["id"]
    current_id = _to_int(group_key[0], -1)
    if current_id < 0:
        return None

    best_id = -1
    best_row = 0
    for row_idx in range(schema["data_start_row"], ws.max_row + 1):
        if row_idx in moving_rows:
            continue
        row_id = _to_int(ws.cell(row_idx, id_col).value, -1)
        if row_id < 0 or row_id >= current_id:
            continue
        if row_id > best_id or (row_id == best_id and row_idx > best_row):
            best_id = row_id
            best_row = row_idx

    if best_row:
        return best_row + 1
    return schema["data_start_row"]


def reorder_payload_rows_openpyxl(ws: Any, sheet_name: str, rows: Iterable[Dict[str, Any]]) -> List[str]:
    schema = SHEET_SCHEMAS[sheet_name]
    key_col = schema["unique_column"]
    max_col = max(schema["field_to_col"].values())
    logs: List[str] = []

    for group_key, desired_keys in _payload_order_groups(sheet_name, rows):
        row_by_key: Dict[str, int] = {}
        desired_set = set(desired_keys)
        for row_idx in range(schema["data_start_row"], ws.max_row + 1):
            key = str(ws.cell(row_idx, key_col).value or "")
            if key in desired_set and key not in row_by_key:
                row_by_key[key] = row_idx

        if any(key not in row_by_key for key in desired_keys):
            continue

        current_start = min(row_by_key.values())
        moving_rows = set(row_by_key.values())
        preferred_start = _payload_group_anchor(ws, sheet_name, group_key, moving_rows) or current_start
        current_keys = [
            str(ws.cell(row_idx, key_col).value or "")
            for row_idx in range(preferred_start, preferred_start + len(desired_keys))
        ]
        if current_keys == desired_keys:
            continue

        insert_start = preferred_start - sum(1 for row_idx in moving_rows if row_idx < preferred_start)

        saved_rows = {
            key: _openpyxl_capture_row(ws, row_by_key[key], max_col)
            for key in desired_keys
        }
        for row_idx in sorted(row_by_key.values(), reverse=True):
            ws.delete_rows(row_idx, 1)
        ws.insert_rows(insert_start, len(desired_keys))
        for offset, key in enumerate(desired_keys):
            _openpyxl_restore_row(ws, insert_start + offset, saved_rows[key])
        logs.append(
            f"{sheet_name}: reorder rows={insert_start}-{insert_start + len(desired_keys) - 1} "
            f"keys={desired_keys[0]}..{desired_keys[-1]}"
        )

    return logs


def summarize_logs(logs: Iterable[str]) -> List[str]:
    summary: Dict[str, Dict[str, Any]] = {}
    reorders: List[str] = []
    dedupes: List[str] = []
    for log in logs:
        action_match = ACTION_LOG_RE.match(log)
        if action_match:
            sheet = action_match.group("sheet")
            item = summary.setdefault(sheet, {"insert": 0, "update": 0, "first_row": None, "last_row": None, "keys": []})
            action = action_match.group("action")
            row = int(action_match.group("row"))
            item[action] += 1
            item["first_row"] = row if item["first_row"] is None else min(item["first_row"], row)
            item["last_row"] = row if item["last_row"] is None else max(item["last_row"], row)
            item["keys"].append(action_match.group("key"))
            continue
        dedupe_match = DEDUPE_LOG_RE.match(log)
        if dedupe_match:
            dedupes.append(
                f"{dedupe_match.group('sheet')} key={dedupe_match.group('key')} remove_rows={dedupe_match.group('rows')}"
            )
            continue
        reorder_match = REORDER_LOG_RE.match(log)
        if reorder_match:
            reorders.append(
                f"{reorder_match.group('sheet')} rows={reorder_match.group('rows')} keys={reorder_match.group('keys')}"
            )

    lines = ["[summary] write plan"]
    for sheet in ("skill", "skill_stage", "buff", "war_paper"):
        item = summary.get(sheet)
        if not item:
            continue
        row_span = f"{item['first_row']}-{item['last_row']}" if item["first_row"] is not None else "-"
        first_key = item["keys"][0] if item["keys"] else "-"
        last_key = item["keys"][-1] if item["keys"] else "-"
        lines.append(
            f"[summary] {sheet}: insert={item['insert']} update={item['update']} rows={row_span} keys={first_key}..{last_key}"
        )
    for item in dedupes:
        lines.append(f"[summary] dedupe: {item}")
    for item in reorders:
        lines.append(f"[summary] reorder: {item}")
    return lines


def verify_payload_rows_openpyxl(
    write_targets: Dict[str, Path],
    payload: Dict[str, Any],
) -> List[str]:
    results: List[str] = []
    workbooks: Dict[str, Any] = {}
    try:
        for workbook_key, workbook_path in write_targets.items():
            workbooks[workbook_key] = load_workbook(workbook_path, read_only=True, data_only=False, keep_links=False)
        for sheet_name, rows in payload["rows"].items():
            if not rows:
                continue
            schema = SHEET_SCHEMAS[sheet_name]
            ws = workbooks[schema["workbook_key"]][schema["sheet_name"]]
            key_col = schema["unique_column"]
            expected = {str(get_unique_value(sheet_name, normalize_row(sheet_name, row))) for row in rows}
            found: Set[str] = set()
            for row_idx in range(schema["data_start_row"], ws.max_row + 1):
                value = str(ws.cell(row_idx, key_col).value or "")
                if value in expected:
                    found.add(value)
            missing = sorted(expected - found)
            if missing:
                results.append(f"[verify-error] {sheet_name}: missing_keys={missing[:10]}")
            else:
                results.append(f"[verify] {sheet_name}: ok keys={len(expected)}")
    finally:
        for wb in workbooks.values():
            close = getattr(wb, "close", None)
            if close:
                close()
    return results


def write_sheet_rows(
    ws: Any,
    sheet_name: str,
    rows: Iterable[Dict[str, Any]],
    dry_run: bool,
    dedupe_existing: bool,
) -> List[str]:
    schema = SHEET_SCHEMAS[sheet_name]
    data_start_row = schema["data_start_row"]
    field_to_col = schema["field_to_col"]
    max_col = max(field_to_col.values())
    signature_map, row_values_by_index, last_row = build_existing_sheet_state(ws, sheet_name, data_start_row)
    logs: List[str] = []
    rows_to_delete: Set[int] = set()

    for raw_row in rows:
        row = normalize_row(sheet_name, raw_row)
        unique_value = get_unique_value(sheet_name, row)
        row_signatures = get_match_signatures(sheet_name, row)
        matched_rows: Set[int] = set()
        for signature in row_signatures:
            matched_rows.update(signature_map.get(signature, []))

        row_idx = min(matched_rows) if matched_rows else None
        action = "update"
        if row_idx is None:
            row_idx = last_row + 1
            last_row = row_idx
            action = "insert"
        elif len(matched_rows) > 1:
            duplicate_rows = sorted(matched_rows - {row_idx})
            if dedupe_existing:
                rows_to_delete.update(duplicate_rows)
                logs.append(
                    f"{sheet_name}: dedupe keep_row={row_idx} remove_rows={duplicate_rows} key={unique_value}"
                )
            else:
                logs.append(
                    f"{sheet_name}: duplicate_detected keep_row={row_idx} duplicate_rows={duplicate_rows} key={unique_value}"
                )

        logs.append(f"{sheet_name}: {action} row={row_idx} key={unique_value}")
        register_row_signatures(signature_map, row_signatures, row_idx)
        if dry_run:
            continue

        output_values = list(row_values_by_index.get(row_idx, [""] * max_col))
        if len(output_values) < max_col:
            output_values.extend([""] * (max_col - len(output_values)))
        for field, col in field_to_col.items():
            if field not in row:
                continue
            output_values[col - 1] = stringify_value(row[field])
        write_row_values(ws, row_idx, output_values)
        row_values_by_index[row_idx] = output_values

    if rows_to_delete:
        for log in delete_rows_desc(ws, rows_to_delete, dry_run):
            logs.append(f"{sheet_name}: {log}")
    return logs


def open_workbook(excel: Any, path: Path, read_only: bool) -> Any:
    return excel.Workbooks.Open(str(path), 0, read_only)


def run_with_openpyxl(args: argparse.Namespace) -> int:
    payload_path = Path(args.payload).expanduser().resolve()
    payload = load_payload(payload_path)
    payload = expand_payload_rows(payload)
    required_keys = required_workbook_keys(payload)
    timings: List[Tuple[str, float]] = []
    total_started_at = time.perf_counter()

    phase_started_at = time.perf_counter()
    targets = resolve_workbook_targets(args, payload, required_keys)
    write_targets = maybe_copy_targets(targets, args.copy_to)
    maybe_backup_targets(write_targets, args.backup_dir, args.dry_run)
    timings.append(("prepare_targets", time.perf_counter() - phase_started_at))

    workbooks: Dict[str, Any] = {}
    touched: Dict[str, bool] = {key: False for key in write_targets}
    all_logs: List[str] = []

    if args.dry_run:
        all_logs = dry_run_with_xlsx_xml(payload, write_targets, args.dedupe_existing, timings)
        timings.append(("total", time.perf_counter() - total_started_at))

        print("payload:", payload_path)
        print("dry_run:", args.dry_run)
        print("dedupe_existing:", args.dedupe_existing)
        print("engine:", args.engine)
        print("scan_mode: xlsx_xml")
        for key, path in write_targets.items():
            print(f"{key}: {path}")
        for log in all_logs:
            print(log)
        for line in summarize_logs(all_logs):
            print(line)
        for label, seconds in timings:
            print(f"[timing] {label}: {seconds:.2f}s")
        if not all_logs:
            print("no rows written")
        return 0

    try:
        for workbook_key, workbook_path in write_targets.items():
            phase_started_at = time.perf_counter()
            workbooks[workbook_key] = load_workbook(
                workbook_path,
                read_only=args.dry_run,
                data_only=False,
                keep_links=False,
            )
            timings.append((f"open_{workbook_key}", time.perf_counter() - phase_started_at))

        for sheet_name, rows in payload["rows"].items():
            if sheet_name not in SHEET_SCHEMAS:
                raise ValueError(f"unsupported sheet name: {sheet_name}")
            if not rows:
                continue
            schema = SHEET_SCHEMAS[sheet_name]
            workbook_key = schema["workbook_key"]
            phase_started_at = time.perf_counter()
            ws = workbooks[workbook_key][schema["sheet_name"]]
            logs = write_sheet_rows_openpyxl(ws, sheet_name, rows, args.dry_run, args.dedupe_existing)
            if not args.dry_run:
                logs.extend(reorder_payload_rows_openpyxl(ws, sheet_name, rows))
            timings.append((f"sheet_{sheet_name}", time.perf_counter() - phase_started_at))
            all_logs.extend(logs)
            if logs:
                touched[workbook_key] = True

        if not args.dry_run:
            for workbook_key, wb in workbooks.items():
                phase_started_at = time.perf_counter()
                if touched[workbook_key]:
                    wb.save(write_targets[workbook_key])
                timings.append((f"save_{workbook_key}", time.perf_counter() - phase_started_at))
    finally:
        for wb in workbooks.values():
            close = getattr(wb, "close", None)
            if close:
                close()

    timings.append(("total", time.perf_counter() - total_started_at))

    print("payload:", payload_path)
    print("dry_run:", args.dry_run)
    print("dedupe_existing:", args.dedupe_existing)
    print("engine:", args.engine)
    for key, path in write_targets.items():
        print(f"{key}: {path}")
    for log in all_logs:
        print(log)
    for line in summarize_logs(all_logs):
        print(line)
    verify_lines = verify_payload_rows_openpyxl(write_targets, payload)
    for line in verify_lines:
        print(line)
    for label, seconds in timings:
        print(f"[timing] {label}: {seconds:.2f}s")
    if not all_logs:
        print("no rows written")
    return 1 if any(line.startswith("[verify-error]") for line in verify_lines) else 0


def run_with_excel_com(args: argparse.Namespace) -> int:
    if win32com is None:
        raise RuntimeError("Excel COM engine requires pywin32. Use --engine openpyxl instead.")

    payload_path = Path(args.payload).expanduser().resolve()
    payload = load_payload(payload_path)
    payload = expand_payload_rows(payload)
    required_keys = required_workbook_keys(payload)
    timings: List[Tuple[str, float]] = []
    total_started_at = time.perf_counter()

    phase_started_at = time.perf_counter()
    targets = resolve_workbook_targets(args, payload, required_keys)
    write_targets = maybe_copy_targets(targets, args.copy_to)
    maybe_backup_targets(write_targets, args.backup_dir, args.dry_run)
    timings.append(("prepare_targets", time.perf_counter() - phase_started_at))

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    excel.Calculation = -4135

    workbooks: Dict[str, Any] = {}
    touched: Dict[str, bool] = {key: False for key in write_targets}
    all_logs: List[str] = []

    try:
        for workbook_key, workbook_path in write_targets.items():
            phase_started_at = time.perf_counter()
            workbooks[workbook_key] = open_workbook(excel, workbook_path, args.dry_run)
            timings.append((f"open_{workbook_key}", time.perf_counter() - phase_started_at))

        for sheet_name, rows in payload["rows"].items():
            if sheet_name not in SHEET_SCHEMAS:
                raise ValueError(f"unsupported sheet name: {sheet_name}")
            if not rows:
                continue
            schema = SHEET_SCHEMAS[sheet_name]
            workbook_key = schema["workbook_key"]
            phase_started_at = time.perf_counter()
            ws = workbooks[workbook_key].Worksheets.Item(schema["sheet_name"])
            logs = write_sheet_rows(ws, sheet_name, rows, args.dry_run, args.dedupe_existing)
            timings.append((f"sheet_{sheet_name}", time.perf_counter() - phase_started_at))
            all_logs.extend(logs)
            if logs:
                touched[workbook_key] = True

        for workbook_key, wb in workbooks.items():
            if args.dry_run:
                wb.Close(False)
            else:
                phase_started_at = time.perf_counter()
                if touched[workbook_key]:
                    wb.Save()
                timings.append((f"save_{workbook_key}", time.perf_counter() - phase_started_at))
                wb.Close(True)
    finally:
        excel.Quit()
    timings.append(("total", time.perf_counter() - total_started_at))

    print("payload:", payload_path)
    print("dry_run:", args.dry_run)
    print("dedupe_existing:", args.dedupe_existing)
    for key, path in write_targets.items():
        print(f"{key}: {path}")
    for log in all_logs:
        print(log)
    for line in summarize_logs(all_logs):
        print(line)
    for label, seconds in timings:
        print(f"[timing] {label}: {seconds:.2f}s")
    if not all_logs:
        print("no rows written")
    return 0


def main() -> int:
    args = parse_args()
    if args.engine == "excel":
        return run_with_excel_com(args)
    return run_with_openpyxl(args)


if __name__ == "__main__":
    sys.exit(main())
